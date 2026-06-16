"""統合テストスクリプト（開発用）"""
import tempfile
import os
import openpyxl
from processor import process_data

csv_content = (
    "ステータス,明細ステータス,商品名,商品グループ,単価（税抜）,請求開始月,請求終了月\n"
    "有効,契約中,商品A,食べログ,5000,2026/01,\n"                     # 残存
    "有効,解約処理完了,商品B,LP,3000,2025/06,2026/03\n"               # ② 除外
    "有効,契約中,商品C,LINE,0,2026/01,\n"                             # ⑤ 除外（0円商材）
    "掲載前キャンセル,契約中,商品D,食べログ,5000,2026/01,\n"           # ⑦ 除外
    "有効,審査不備・キャンセル,商品E,LP,3000,2026/01,\n"              # ③ 除外
    "有効,契約中,,その他,5000,2026/01,\n"                             # ⑧ 除外（商品名空白）
    "有効,契約中,初期,食べログ,0,2026/01,2025/12\n"                   # ④ 除外
    "有効,契約中,商品F,食べログ,5000,2026/01,2026/04\n"               # 残存（終了月>=当月）
    "有効,契約中,商品G,LINE,8000,2026/01,2026/03\n"                   # ⑥ 除外（先月終了）
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "単価0円商材"
ws.append(["商品名", "備考"])
ws.append(["商品C", "0円商材テスト"])
wb.create_sheet("データの絞り方").append(["フィルタ条件"])

tmp_excel = tempfile.mktemp(suffix=".xlsx")
tmp_csv = tempfile.mktemp(suffix=".csv")
wb.save(tmp_excel)
with open(tmp_csv, "w", encoding="utf-8-sig") as f:
    f.write(csv_content)

result = process_data(
    csv_path=tmp_csv,
    prev_excel_path=tmp_excel,
    target_year=2026,
    target_month=4,
    progress_callback=print,
)

print()
print(f"errors   : {result.errors}")
print(f"warnings : {result.warnings}")
print(f"total    : {result.total_count}")
print(f"excluded : {result.excluded_count}")
print(f"output   : {result.output_count}")
print(f"sheets   : {result.workbook.sheetnames if result.workbook else None}")
print(f"tab_counts: {result.tab_counts}")

ws_moto = result.workbook["moto"]
headers = [c.value for c in ws_moto[1]]
ex_num_col = headers.index("除外番号")
shohin_col = headers.index("商品名")
print()
print("除外番号一覧:")
for row in ws_moto.iter_rows(min_row=2, values_only=True):
    num = row[ex_num_col]
    name = row[shohin_col]
    if num:
        print(f"  商品名={name}  除外番号={num}")

os.unlink(tmp_csv)
os.unlink(tmp_excel)
print("\nテスト完了")
