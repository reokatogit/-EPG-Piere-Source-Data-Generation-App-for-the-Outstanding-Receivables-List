"""xlsx ファイル入力のテスト"""
import tempfile
import os
import openpyxl
from processor import process_data

# テスト用 Excel ファイル（入力）を作成
wb_input = openpyxl.Workbook()
ws_input = wb_input.active
ws_input.title = "契約"

# ヘッダ行
ws_input.append(["ステータス", "明細ステータス", "商品名", "商品グループ", "単価（税抜）", "請求開始月", "請求終了月"])

# データ行
ws_input.append(["有効", "契約中", "商品A", "食べログ", "5000", "2026/01", ""])
ws_input.append(["有効", "解約処理完了", "商品B", "LP", "3000", "2025/06", "2026/03"])
ws_input.append(["有効", "契約中", "商品C", "LINE", "0", "2026/01", ""])
ws_input.append(["掲載前キャンセル", "契約中", "商品D", "食べログ", "5000", "2026/01", ""])
ws_input.append(["有効", "審査不備・キャンセル", "商品E", "LP", "3000", "2026/01", ""])

# テスト用 Excel ファイル（前月分）を作成
wb_prev = openpyxl.Workbook()
ws_prev = wb_prev.active
ws_prev.title = "単価0円商材"
ws_prev.append(["商品名", "備考"])
ws_prev.append(["商品C", "0円商材テスト"])
wb_prev.create_sheet("データの絞り方").append(["フィルタ条件"])

# 一時ファイルに保存
tmp_input_xlsx = tempfile.mktemp(suffix=".xlsx")
tmp_prev_xlsx = tempfile.mktemp(suffix=".xlsx")

wb_input.save(tmp_input_xlsx)
wb_prev.save(tmp_prev_xlsx)

print(f"入力ファイル (xlsx): {tmp_input_xlsx}")
print(f"前月ファイル (xlsx): {tmp_prev_xlsx}")
print()

# 処理実行
result = process_data(
    csv_path=tmp_input_xlsx,  # xlsx ファイルを入力
    prev_excel_path=tmp_prev_xlsx,
    target_year=2026,
    target_month=4,
    progress_callback=print,
)

print()
print(f"errors   : {result.errors}")
print(f"warnings : {result.warnings}")
print(f"total    : {result.total_count}")
print(f"excluded : {result.excluded_count}")
print(f"output   : {result.output_count}")

if result.errors:
    print("\n【エラー】Excel 入力に対応していません")
else:
    print("\n【成功】Excel 入力ファイルを正常に処理できました!")

os.unlink(tmp_input_xlsx)
os.unlink(tmp_prev_xlsx)
print("テスト完了")
