"""
processor.py - 業務ロジック（除外判定・シート生成・Excel出力）

処理フロー:
  1. CSV 読込 → DataFrame (moto)
  2. 単価0円商材の読込（以下いずれか）
       a. 単価0円商材ファイル（単価0円商材_データ絞り方.xlsx 等）を指定した場合 → そちらを優先
       b. 前月分完成Excel を指定した場合 → 「単価0円商材」「データの絞り方」シートを読込
  3. 除外判定 ②〜⑧ を順番に適用
  4. in-memory で openpyxl Workbook を生成（シート一式）
  5. GUI が警告確認後に save_result() を呼んで保存
"""

from __future__ import annotations

import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Callable

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import (
    EXTRA_COLS_DISPLAY,
    EXTRA_COLS_INTERNAL,
    INTERNAL_TO_DISPLAY,
    find_column,
    month_to_int,
    normalize_col_name,
    parse_month,
    prev_month,
)
from fast_xlsx_reader import is_large_xlsx, read_large_xlsx

logger = logging.getLogger(__name__)

# ==================== 定数 ====================

REQUIRED_COLUMNS = [
    "ステータス",
    "明細ステータス",
    "商品名",
    "商品グループ",
    "単価（税抜）",
    "請求開始月",
    "請求終了月",
]


# ==================== データクラス ====================

@dataclass
class ProcessingResult:
    total_count: int = 0
    excluded_count: int = 0
    output_count: int = 0
    output_file: str = ""
    log_file: str = ""
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    workbook: Workbook | None = None
    tab_counts: dict[str, int] = field(default_factory=dict)
    # 保存時に使う情報
    _target_year: int = 0
    _target_month: int = 0
    _csv_path: str = ""
    _zero_price_xlsx_path: str = ""
    _prev_excel_path: str = ""


# ==================== CSV・Excel 読込 ====================

def read_csv(path: str) -> pd.DataFrame:
    """複数エンコーディングを試みて CSV を読み込む。"""
    for enc in ("utf-8-sig", "utf-8", "cp932", "shift-jis", "shift_jis"):
        try:
            df = pd.read_csv(path, encoding=enc, dtype=str)
            df.columns = [normalize_col_name(c) for c in df.columns]
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            raise ValueError(f"CSV読込エラー: {e}") from e
    raise ValueError("CSVファイルの文字コードを認識できませんでした（UTF-8/Shift-JIS を試みましたが失敗）。")


def read_excel(path: str) -> pd.DataFrame:
    """Excel ファイル（xlsx/xlsm）から必須列を含むシートを自動判別して読み込む。"""
    try:
        # openpyxl read_only でヘッダだけ高速チェック（全データを読まない）
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        target_sheet = wb.sheetnames[0]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                cols = [normalize_col_name(str(c)) for c in header_row if c is not None]
                if all(find_column(cols, req_col) is not None for req_col in REQUIRED_COLUMNS):
                    target_sheet = sheet_name
                    break
        wb.close()

        # 該当シートだけ pandas で読み込み
        logger.info(f"シート'{target_sheet}'を読み込みます")
        df = pd.read_excel(path, sheet_name=target_sheet, dtype=str, engine="openpyxl")
        df.columns = [normalize_col_name(c) for c in df.columns]
        return df
    except Exception as e:
        raise ValueError(f"Excel読込エラー: {e}") from e


def read_input_file(path: str, progress_callback=None) -> pd.DataFrame:
    """CSV または Excel ファイルを自動判別して読み込む。"""
    _, ext = os.path.splitext(path)
    ext = ext.lower()

    if ext in (".csv",):
        return read_csv(path)
    elif ext in (".xlsx", ".xlsm"):
        if is_large_xlsx(path):
            return read_large_xlsx(
                path, REQUIRED_COLUMNS, find_column, normalize_col_name,
                progress=progress_callback,
            )
        return read_excel(path)
    else:
        raise ValueError(f"サポートされていないファイル形式です: {ext}（CSV または Excel ファイルを選択してください）")


# ==================== 除外判定 ====================


def apply_exclusions(
    df: pd.DataFrame,
    col_map: dict[str, str],
    current_month: tuple[int, int],
    zero_price_products: set[str],
    warnings: list[str],
) -> pd.DataFrame:
    """
    除外判定ルール ②〜⑧ を順番に適用する。
    各除外対象行に「除外」=「除外」、「除外番号」= ②〜⑧ を設定する。
    """
    current_int = month_to_int(current_month)
    prev_mo_int = month_to_int(prev_month(current_month))

    df = df.copy()
    for col in EXTRA_COLS_INTERNAL:
        df[col] = ""

    meisai_col = col_map.get("明細ステータス")
    shohin_col = col_map.get("商品名")
    group_col = col_map.get("商品グループ")
    tanka_col = col_map.get("単価（税抜）")
    seikyu_end_col = col_map.get("請求終了月")
    status_col = col_map.get("ステータス")

    # 各列を Series として取得（欠損は空文字）
    def get_col(col: str | None) -> pd.Series:
        if col is None:
            return pd.Series("", index=df.index)
        return df[col].fillna("").astype(str).str.strip()

    meisai = get_col(meisai_col)
    shohin = get_col(shohin_col)
    tanka_raw = get_col(tanka_col)
    seikyu_end_raw = get_col(seikyu_end_col)
    status = get_col(status_col)

    # 請求終了月を一括解析して整数化
    seikyu_end_int = seikyu_end_raw.map(lambda v: month_to_int(parse_month(v)) if v not in ("", "nan", "None") else None)

    # 日付解析失敗カウント（空でなく解析できなかった件数）
    date_parse_failures = int(
        seikyu_end_raw[seikyu_end_raw.str.strip().ne("") &
                       seikyu_end_raw.str.lower().ne("nan") &
                       seikyu_end_raw.str.lower().ne("none")].map(
            lambda v: parse_month(v) is None
        ).sum()
    )

    # 単価を数値化
    def parse_tanka(v: str) -> float | None:
        try:
            return float(v.replace(",", "").replace("，", "").replace("¥", "").replace("\\", ""))
        except (ValueError, TypeError, AttributeError):
            return None

    tanka_val = tanka_raw.map(parse_tanka)

    # 単価0円商材フラグ（商品名を正規化して照合）
    shohin_norm = shohin.map(normalize_col_name)
    is_zero_product = shohin_norm.isin(zero_price_products)
    df.loc[is_zero_product, "単価0円案件判別"] = "○"

    # ── 除外判定をベクトルで適用（優先順位順） ──────────
    excluded = pd.Series(False, index=df.index)

    def mark(mask: pd.Series, number: str, reason: str) -> None:
        """まだ除外されていない行に除外フラグを立てる。"""
        new_mask = mask & ~excluded
        df.loc[new_mask, "除外"] = "除外"
        df.loc[new_mask, "除外番号"] = number
        df.loc[new_mask, "除外理由"] = reason
        excluded.update(excluded | new_mask)

    # ② 明細ステータス = 解約処理完了（請求終了月が当月以降のものは除外しない）
    mask_kaiyaku = meisai.eq("解約処理完了")
    mask_kaiyaku_keep = seikyu_end_int.ge(current_int).fillna(False)
    mark(mask_kaiyaku & ~mask_kaiyaku_keep, "②", "解約処理完了")

    # ③ 明細ステータス = 審査不備・キャンセル
    mark(meisai.eq("審査不備・キャンセル"), "③", "審査不備・キャンセル")

    # ④ 商品名 = 初期 かつ 請求終了月が空白または過去
    mask_shoki_keep = seikyu_end_int.ge(current_int).fillna(False)
    mark(shohin.eq("初期") & ~mask_shoki_keep, "④", "初期商品（請求終了済）")

    # ⑤ 単価0円商材
    mark(is_zero_product & tanka_val.eq(0.0), "⑤", "単価0円商材")

    # ⑥ 請求終了月が先月以前
    mark(seikyu_end_int.le(prev_mo_int).fillna(False), "⑥", "請求終了月が先月以前")

    # ⑦ ステータス = 掲載前キャンセル
    mark(status.eq("掲載前キャンセル"), "⑦", "掲載前キャンセル")

    # ⑧ 商品名が空白
    mark(shohin.eq(""), "⑧", "商品名空白")

    # 警告追記
    if date_parse_failures > 0:
        warnings.append(f"請求終了月の日付解釈に失敗した行が {date_parse_failures} 件あります（除外判定をスキップ）。")

    return df


# ==================== Excel シート書き込み ====================

def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    """DataFrame を openpyxl ワークシートに書き込む（ヘッダ含む）。"""
    headers = [INTERNAL_TO_DISPLAY.get(c, c) for c in df.columns]
    ws.append(headers)
    # fillna だけで NaN/None を処理（データは既に str 型なので astype 不要）
    rows = df.fillna("").values.tolist()
    for row in rows:
        ws.append(row)


def _copy_rows_to_sheet(target_wb: Workbook, sheet_name: str, rows: list[tuple]):
    """行データリストを新ワークブックのシートに書き込む。"""
    ws = target_wb.create_sheet(title=sheet_name)
    for row in rows:
        ws.append([("" if v is None else v) for v in row])


# ==================== メイン処理（データ加工フェーズ） ====================

def process_data(
    csv_path: str,
    target_year: int,
    target_month: int,
    zero_price_xlsx_path: str = "",
    prev_excel_path: str = "",
    progress_callback: Callable[[str], None] | None = None,
) -> ProcessingResult:
    """
    データを読み込み・加工して in-memory Workbook を生成して返す。
    zero_price_xlsx_path と prev_excel_path はどちらか一方（または両方）を指定する。
    両方指定時は zero_price_xlsx_path を優先して単価0円商材を読込む。
    保存は行わない（GUI が警告確認後に save_result() を呼ぶ）。
    """
    result = ProcessingResult(
        _target_year=target_year,
        _target_month=target_month,
        _csv_path=csv_path,
        _zero_price_xlsx_path=zero_price_xlsx_path,
        _prev_excel_path=prev_excel_path,
    )

    def progress(msg: str) -> None:
        logger.info(msg)
        if progress_callback:
            progress_callback(msg)

    # ── CSV 読込 ─────────────────────────────────
    progress("今月分ファイル読込中...")
    try:
        df = read_input_file(csv_path, progress_callback=progress_callback)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    result.total_count = len(df)
    progress(f"ファイル読込完了: {result.total_count} 件")

    # 必須列チェック
    col_map: dict[str, str] = {}
    missing_cols: list[str] = []
    for req_col in REQUIRED_COLUMNS:
        found = find_column(df.columns, req_col)
        if found:
            col_map[req_col] = found
        else:
            missing_cols.append(req_col)
    if missing_cols:
        result.errors.append(f"必須列が見つかりません: {', '.join(missing_cols)}")
        return result

    # ── 単価0円商材・データの絞り方 読込 ─────────────────
    zero_price_products: set[str] = set()
    zero_price_rows: list[tuple] = []
    prev_filter_rows: list[tuple] = []

    def _load_zero_price_sheet(wb_obj) -> None:
        """openpyxl Workbook から単価0円商材シートを読み込んで zero_price_* に格納。"""
        if "単価0円商材" not in wb_obj.sheetnames:
            result.warnings.append("指定ファイルに「単価0円商材」シートが見つかりません。除外⑤の判定をスキップします。")
            return
        ws = wb_obj["単価0円商材"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        shohin_col_idx = 0
        if header_row:
            zero_price_rows.append(header_row)
            for i, h in enumerate(header_row):
                if h and "商品" in str(h):
                    shohin_col_idx = i
                    break
        for row in ws.iter_rows(min_row=2, values_only=True):
            zero_price_rows.append(row)
            if row and len(row) > shohin_col_idx and row[shohin_col_idx]:
                val = str(row[shohin_col_idx]).strip()
                if val:
                    zero_price_products.add(normalize_col_name(val))

    if zero_price_xlsx_path:
        # ── 専用ファイル（単価0円商材_データ絞り方.xlsx 等）から読込 ──
        progress("単価0円商材ファイル読込中...")
        try:
            zp_wb = openpyxl.load_workbook(zero_price_xlsx_path, read_only=True, data_only=True)
            _load_zero_price_sheet(zp_wb)
            # データの絞り方も同ファイルにあれば読込
            if "データの絞り方" in zp_wb.sheetnames:
                for row in zp_wb["データの絞り方"].iter_rows(values_only=True):
                    prev_filter_rows.append(row)
            zp_wb.close()
            progress(f"単価0円商材: {len(zero_price_products)} 件読込")
        except Exception as e:
            result.errors.append(f"単価0円商材ファイルの読込に失敗しました: {e}")
            return result

        # 前月分Excelが同時に指定されていればデータの絞り方を上書き（優先）
        if prev_excel_path:
            try:
                prev_wb = openpyxl.load_workbook(prev_excel_path, read_only=True, data_only=True)
                if "データの絞り方" in prev_wb.sheetnames:
                    prev_filter_rows.clear()
                    for row in prev_wb["データの絞り方"].iter_rows(values_only=True):
                        prev_filter_rows.append(row)
                prev_wb.close()
            except Exception:
                pass  # データの絞り方は補助情報なのでエラーでも続行

    elif prev_excel_path:
        # ── 前月分完成Excel から読込（旧来方式）──
        progress("前月分Excel読込中...")
        try:
            prev_wb = openpyxl.load_workbook(prev_excel_path, read_only=True, data_only=True)
        except Exception as e:
            result.errors.append(f"前月分完成Excelの読込に失敗しました: {e}")
            return result
        _load_zero_price_sheet(prev_wb)
        progress(f"単価0円商材: {len(zero_price_products)} 件読込")
        if "データの絞り方" in prev_wb.sheetnames:
            for row in prev_wb["データの絞り方"].iter_rows(values_only=True):
                prev_filter_rows.append(row)
            progress("データの絞り方シート読込完了")
        else:
            result.warnings.append("前月分Excelに「データの絞り方」シートが見つかりません。空シートを作成します。")
        prev_wb.close()

    else:
        result.errors.append("単価0円商材ファイルまたは前月分完成Excelのいずれかを指定してください。")
        return result

    # ── 除外判定 ─────────────────────────────────
    progress("除外判定中...")
    current_month = (target_year, target_month)
    df = apply_exclusions(df, col_map, current_month, zero_price_products, result.warnings)

    result.excluded_count = int((df["除外"] == "除外").sum())
    result.output_count = result.total_count - result.excluded_count
    progress(f"除外判定完了: 除外 {result.excluded_count} 件 / 残 {result.output_count} 件")

    # ── Excel ワークブック生成 ────────────────────
    # シート順序（先月準拠）: Sheet5, 単価0円商材, データの絞り方, moto, 除外レコード除外済,
    #   その他商材, LINE, LP, 食べログ
    progress("Excelワークブック生成中...")
    wb = Workbook()
    wb.remove(wb.active)

    # ⓪ Sheet5（空シート・先月互換）
    wb.create_sheet("Sheet5")

    group_col_name = col_map.get("商品グループ")

    # ① 単価0円商材シート
    if zero_price_rows:
        _copy_rows_to_sheet(wb, "単価0円商材", zero_price_rows)
    else:
        wb.create_sheet("単価0円商材")

    # ② データの絞り方シート（前月分Excelから読込した場合のみ）
    if prev_filter_rows:
        _copy_rows_to_sheet(wb, "データの絞り方", prev_filter_rows)

    # ③ moto シート（全データ + 追加7列）
    ws_moto = wb.create_sheet("moto")
    _write_df_to_sheet(ws_moto, df)
    progress("motoシート作成完了")

    # 除外後データ（除外 != "除外" の行）
    df_valid = df[df["除外"] != "除外"].copy()
    df_valid["契約有無"] = "契約有"

    # ④ 除外レコード除外済シート
    ws_valid = wb.create_sheet("除外レコード除外済")
    _write_df_to_sheet(ws_valid, df_valid)

    # 商品グループ別シート（除外後データから振り分け）
    if group_col_name:
        mask_tabelog = df_valid[group_col_name].str.contains("食べログ", na=False)
        mask_lp = df_valid[group_col_name].str.contains("LP", na=False) & ~mask_tabelog
        mask_line = df_valid[group_col_name].str.contains("LINE", na=False) & ~mask_tabelog & ~mask_lp
        mask_other = ~mask_tabelog & ~mask_lp & ~mask_line

        # ⑤ その他商材シート
        df_other = df_valid[mask_other].copy()
        df_other["契約有無"] = df_other[group_col_name].fillna("").astype(str)
        ws_other = wb.create_sheet("その他商材")
        _write_df_to_sheet(ws_other, df_other)

        # ⑥ LINEシート
        df_line = df_valid[mask_line].copy()
        df_line["契約有無"] = "LINE 有効"
        ws_line = wb.create_sheet("LINE")
        _write_df_to_sheet(ws_line, df_line)

        # ⑦ LPシート
        df_lp = df_valid[mask_lp].copy()
        df_lp["契約有無"] = "LP 有効"
        ws_lp = wb.create_sheet("LP")
        _write_df_to_sheet(ws_lp, df_lp)

        # ⑧ 食べログシート
        df_tabelog = df_valid[mask_tabelog].copy()
        df_tabelog["契約有無"] = "食べログ有効"
        ws_tabelog = wb.create_sheet("食べログ")
        _write_df_to_sheet(ws_tabelog, df_tabelog)

        result.tab_counts = {
            "食べログ": len(df_tabelog),
            "LP": len(df_lp),
            "LINE": len(df_line),
            "その他商材": len(df_other),
        }
        progress(
            f"商材別シート作成完了 ("
            f"食べログ: {len(df_tabelog)} 件 / "
            f"LP: {len(df_lp)} 件 / "
            f"LINE: {len(df_line)} 件 / "
            f"その他: {len(df_other)} 件)"
        )
    else:
        # 商品グループ列が無い場合（通常ありえないが保護）
        result.warnings.append("商品グループ列が見つからないため商材別シートを作成できません。")
        for sheet_name in ("食べログ", "LP", "LINE", "その他商材"):
            wb.create_sheet(sheet_name)

    result.workbook = wb
    progress("Excelワークブック生成完了")
    return result


# ==================== 保存フェーズ ====================

def save_result(result: ProcessingResult, output_folder: str, output_name: str = "") -> str | None:
    """
    生成済み Workbook をファイルに保存する。
    成功すれば None、エラーがあればエラーメッセージ文字列を返す。
    """
    if result.workbook is None:
        return "保存対象のワークブックがありません。"

    if not output_name:
        target_year = result._target_year
        target_month = result._target_month
        output_name = f"【Piere契約有効案件{target_year}{target_month:02d}まで】nv_contract_v_SJIS"

    filename = f"{output_name}.xlsx"
    output_path = os.path.join(output_folder, filename)

    try:
        os.makedirs(output_folder, exist_ok=True)
        result.workbook.save(output_path)
    except Exception as e:
        return f"Excelファイルの保存に失敗しました: {e}"

    result.output_file = output_path
    return None


def _write_log(result: ProcessingResult, log_path: str) -> None:
    """処理ログをテキストファイルに書き出す。"""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"処理日時       : {now_str}",
        f"処理対象年月   : {result._target_year}年{result._target_month}月",
        f"入力CSV            : {result._csv_path}",
        f"単価0円商材ファイル: {result._zero_price_xlsx_path or '（未指定）'}",
        f"前月分Excel        : {result._prev_excel_path or '（未指定）'}",
        f"出力ファイル       : {result.output_file}",
        "",
        "--- 処理結果 ---",
        f"読込件数 : {result.total_count} 件",
        f"除外件数 : {result.excluded_count} 件",
        f"出力件数 : {result.output_count} 件",
    ]
    if result.tab_counts:
        lines.append("")
        lines.append("--- 商材別件数 ---")
        for sheet, cnt in result.tab_counts.items():
            lines.append(f"  {sheet}: {cnt} 件")
    if result.warnings:
        lines.append("")
        lines.append("--- 警告 ---")
        for w in result.warnings:
            lines.append(f"  [警告] {w}")

    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
